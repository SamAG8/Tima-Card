"""
Excel export service using openpyxl.
Generates payroll and time entry reports with budget code breakdown.
"""
import io
from datetime import date
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


PRIMARY_COLOR  = "1E3A5F"
SUCCESS_COLOR  = "166534"
ACCENT_COLOR   = "7C3AED"
HEADER_FG      = "FFFFFF"
ALT_ROW_COLOR  = "F0F4F8"
DIVISION_COLOR = "E8F0FE"


def _style_header_row(ws, row: int, col_count: int, bg: str = PRIMARY_COLOR):
    fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    font = Font(bold=True, color=HEADER_FG)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_data_row(ws, row: int, col_count: int, alternate: bool):
    if alternate:
        fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid")
        for col in range(1, col_count + 1):
            ws.cell(row=row, column=col).fill = fill


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generate_payroll_excel(
    rows: list[dict],
    start_date: date,
    end_date: date,
    company_name: str = "Company",
    worker_names: Optional[dict] = None,
    project_names: Optional[dict] = None,
) -> bytes:
    wb = openpyxl.Workbook()

    # -------------------------------------------------------------------------
    # Sheet 1: Time Entries Detail
    # -------------------------------------------------------------------------
    ws_detail = wb.active
    ws_detail.title = "Time Entries Detail"

    ws_detail.append([f"Payroll Report — {company_name}"])
    ws_detail.append([f"{start_date} to {end_date}"])
    ws_detail.append([])

    headers = [
        "Date", "Worker", "Project",
        "Division", "Category", "Budget Code", "Budget Code Name",
        "Description",
        "Clock In", "Clock Out",
        "Hours Worked", "Hourly Rate", "Currency", "Total Cost",
    ]
    ws_detail.append(headers)
    _style_header_row(ws_detail, 4, len(headers))

    for i, row in enumerate(rows):
        worker  = (worker_names  or {}).get(row["user_id"],    row["user_id"])
        project = (project_names or {}).get(row["project_id"], row["project_id"])
        desc = (row.get("description") or "").strip()
        if not desc:
            desc = "Nothing written"
        ws_detail.append([
            row["work_date"],
            worker,
            project,
            row.get("division")         or "—",
            row.get("category")         or "—",
            row.get("budget_code")      or "—",
            row.get("budget_code_name") or "—",
            desc,
            row.get("clock_in",  ""),
            row.get("clock_out", ""),
            round(row["hours_worked"], 2),
            row["hourly_rate"],
            row["currency"],
            row["total_cost"],
        ])
        _style_data_row(ws_detail, 5 + i, len(headers), i % 2 == 1)

    _auto_width(ws_detail)

    # -------------------------------------------------------------------------
    # Sheet 2: Summary by Worker
    # -------------------------------------------------------------------------
    ws_worker = wb.create_sheet("Summary by Worker")
    ws_worker.append(["Worker", "Total Hours", "Total Cost", "Currency", "Entries"])
    _style_header_row(ws_worker, 1, 5)

    agg: dict[str, dict] = {}
    for row in rows:
        uid = row["user_id"]
        if uid not in agg:
            agg[uid] = {"hours": 0.0, "cost": 0.0, "currency": row["currency"], "entries": 0}
        agg[uid]["hours"]   += row["hours_worked"]
        agg[uid]["cost"]    += row["total_cost"] or 0.0
        agg[uid]["entries"] += 1

    for i, (uid, data) in enumerate(agg.items()):
        worker = (worker_names or {}).get(uid, uid)
        ws_worker.append([
            worker,
            round(data["hours"], 2),
            round(data["cost"],  2),
            data["currency"],
            data["entries"],
        ])
        _style_data_row(ws_worker, 2 + i, 5, i % 2 == 1)

    _auto_width(ws_worker)

    # -------------------------------------------------------------------------
    # Sheet 3: Summary by Project
    # -------------------------------------------------------------------------
    ws_proj = wb.create_sheet("Summary by Project")
    ws_proj.append(["Project", "Total Hours", "Total Cost", "Currency", "Entries"])
    _style_header_row(ws_proj, 1, 5)

    proj_agg: dict[str, dict] = {}
    for row in rows:
        pid = row["project_id"]
        if pid not in proj_agg:
            proj_agg[pid] = {"hours": 0.0, "cost": 0.0, "currency": row["currency"], "entries": 0}
        proj_agg[pid]["hours"]   += row["hours_worked"]
        proj_agg[pid]["cost"]    += row["total_cost"] or 0.0
        proj_agg[pid]["entries"] += 1

    for i, (pid, data) in enumerate(proj_agg.items()):
        project = (project_names or {}).get(pid, pid)
        ws_proj.append([
            project,
            round(data["hours"], 2),
            round(data["cost"],  2),
            data["currency"],
            data["entries"],
        ])
        _style_data_row(ws_proj, 2 + i, 5, i % 2 == 1)

    _auto_width(ws_proj)

    # -------------------------------------------------------------------------
    # Sheet 4: By Budget Code
    # -------------------------------------------------------------------------
    ws_bc = wb.create_sheet("By Budget Code")
    bc_headers = ["Division", "Category", "Budget Code", "Budget Code Name", "Total Hours", "Total Cost", "Currency", "Entries"]
    ws_bc.append(bc_headers)
    _style_header_row(ws_bc, 1, len(bc_headers), bg=SUCCESS_COLOR)

    bc_agg: dict[str, dict] = {}
    for row in rows:
        key = row.get("budget_code_id") or "__untagged__"
        if key not in bc_agg:
            bc_agg[key] = {
                "division":         row.get("division")         or "Untagged",
                "category":         row.get("category")         or "—",
                "budget_code":      row.get("budget_code")      or "—",
                "budget_code_name": row.get("budget_code_name") or "No Budget Code",
                "hours":  0.0,
                "cost":   0.0,
                "currency": row["currency"],
                "entries": 0,
            }
        bc_agg[key]["hours"]   += row["hours_worked"]
        bc_agg[key]["cost"]    += row["total_cost"] or 0.0
        bc_agg[key]["entries"] += 1

    sorted_bc = sorted(bc_agg.values(), key=lambda x: x["hours"], reverse=True)
    for i, data in enumerate(sorted_bc):
        ws_bc.append([
            data["division"],
            data["category"],
            data["budget_code"],
            data["budget_code_name"],
            round(data["hours"],  2),
            round(data["cost"],   2),
            data["currency"],
            data["entries"],
        ])
        _style_data_row(ws_bc, 2 + i, len(bc_headers), i % 2 == 1)

    _auto_width(ws_bc)

    # -------------------------------------------------------------------------
    # Sheet 5: By Division
    # -------------------------------------------------------------------------
    ws_div = wb.create_sheet("By Division")
    div_headers = ["Division", "Total Hours", "Total Cost", "Currency", "Entries"]
    ws_div.append(div_headers)
    _style_header_row(ws_div, 1, len(div_headers), bg=ACCENT_COLOR)

    div_agg: dict[str, dict] = {}
    for row in rows:
        key = row.get("division") or "Untagged"
        if key not in div_agg:
            div_agg[key] = {"hours": 0.0, "cost": 0.0, "currency": row["currency"], "entries": 0}
        div_agg[key]["hours"]   += row["hours_worked"]
        div_agg[key]["cost"]    += row["total_cost"] or 0.0
        div_agg[key]["entries"] += 1

    sorted_div = sorted(div_agg.items(), key=lambda x: x[1]["hours"], reverse=True)
    for i, (div_name, data) in enumerate(sorted_div):
        ws_div.append([
            div_name,
            round(data["hours"], 2),
            round(data["cost"],  2),
            data["currency"],
            data["entries"],
        ])
        _style_data_row(ws_div, 2 + i, len(div_headers), i % 2 == 1)

    _auto_width(ws_div)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
